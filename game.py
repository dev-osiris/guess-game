#! python3
import random
from openpyxl import *
low_num = random.randint(1, 5)
high_num = random.randint(18, 22)
number = random.randint(low_num, high_num)
wb = load_workbook('guess_game.xlsx')
sheet = wb['Sheet1']
sheet['A1'] = 'random limit'
sheet['B1'] = 'random number'
sheet['C1'] = 'attempts taken'
guessnumber = ''
# uncomment to show answer beforehand
# print('answer: ' + str(number))
try:
    guessnumber = int(input(f'Guess a num between {low_num} and {high_num}' +
                            ' '*25 + 'lives left-'+5 * "\U0001f600" + '\n'))
except ValueError:
    print('please enter only a number.')
    exit()
feed = True
i = 0
j = 4
while feed:
    if i >= 4 and guessnumber != number:
        print(f'GAME OVER!,actual number was {number}.')
        break
    if guessnumber < number and low_num <= guessnumber <= high_num:
        print(f'your guess is too low.\nTake a guess.', end='')
        print(' '*40 + 'lives left-' + i*' ' + j * "\U0001f600")
        try:
            guessnumber = int(input())
        except ValueError:
            print('please enter a number. Try again')
            exit()
    elif guessnumber > number and low_num <= guessnumber <= high_num:
        print(f'your guess is too high.\nTake a guess.', end='')
        print(' '*40 + 'lives left-' + i*' ' + j * "\U0001f600")
        try:
            guessnumber = int(input())
        except ValueError:
            print('please enter a number. Try again')
            exit()
    elif guessnumber == number:
        feed = 0
        if i == 0:
            print('good job! you guessed number right in 1 guess.')
        else:
            print(f'good job! you guessed number right in {i+1} guesses.')
        break
    else:
        print(f'I said between {low_num} to {high_num}!')
        exit()
    i += 1
    j -= 1
row1 = sheet.max_row + 1
sheet.cell(row=row1, column=1).value = str(low_num) + ' - ' + str(high_num)
sheet.cell(row=row1, column=2).value = number
if i >= 4 and guessnumber != number:
    sheet.cell(row=row1, column=3).value = '       GAME OVER'
else:
    sheet.cell(row=row1, column=3).value = i+1
try:
    wb.save('guess_game.xlsx')
except PermissionError:
    print('\nERROR : Cannot save file as it is already opened in Excel.')
    exit()
wb.close()
